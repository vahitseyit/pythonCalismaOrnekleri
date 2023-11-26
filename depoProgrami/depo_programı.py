from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def kayit_ekle():
    while True:
        # Kullanıcıdan TC Kimlik No'yu al
        tc_no = input("TC Kimlik No (Çıkmak için q): ")

        # Programdan çık
        if tc_no.lower() == "q":
            break

        # Dosya varsa mevcut verileri oku
        wb = load_workbook(dosya_yolu) if dosya_yolu in os.listdir() else Workbook()
        ws = wb.active
        header = ["TC Kimlik No", "Ad", "Soyad", "Ürün Adı", "Tarih"]

        # Eğer TC Kimlik No zaten kayıtlıysa uyarı ver ve tekrar iste
        if tc_no in [str(row[0]) for row in ws.iter_rows(min_row=2, min_col=1, values_only=True) if row[0]]:
            print("Bu TC kimlik numarası zaten kaydedilmiş.")
            kayitlar = [row for row in ws.iter_rows(min_row=2, min_col=1, values_only=True) if row[0] == tc_no]
            print(f"Toplam {len(kayitlar)} kayıt mevcut.")
            devam_et = input("Kayıt etmeye devam etmek istiyor musunuz? (E/H): ")
            if devam_et.lower() != "e":
                continue

        ad = input("Ad: ")
        soyad = input("Soyad: ")
        urun_ad = input("Ürün Adı: ")
        tarih = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Yeni kaydı veri çerçevesine ekle
        new_row = [tc_no, ad, soyad, urun_ad, tarih]
        ws.append(new_row)

        # Verileri dosyaya kaydet
        wb.save(dosya_yolu)

        print(f"{tc_no} TC kimlik numarası ile kayıt eklendi.\n")
        os.system('cls' if os.name == 'nt' else 'clear')






def kayit_sil():
    while True:
        # Kullanıcıdan TC Kimlik No'yu al
        tc_no = input("TC Kimlik No (Çıkmak için q): ")
        # Programdan çık
        if tc_no.lower() == "q":
            break

        # Dosya varsa mevcut verileri oku
        wb = load_workbook(dosya_yolu) if dosya_yolu in os.listdir() else Workbook()
        ws = wb.active

        # TC Kimlik No'ya göre kaydı bul
        kayitlar = [row for row in ws.iter_rows(min_row=2, min_col=1, values_only=True) if row[0] == tc_no]

        # Kayıt bulunamadıysa uyarı ver ve tekrar iste
        if not kayitlar:
            print("Bu TC kimlik numarasına sahip bir kayıt bulunamadı.")
            continue

        # Birden fazla kayıt varsa kullanıcıdan hangisinin silineceğini seçmesini iste
        if len(kayitlar) > 1:
            print(f"Toplam {len(kayitlar)} kayıt bulundu:")
            for i, kayit in enumerate(kayitlar, 1):
                print(f"{i}. Kayıt:")
                print(f"\tAdı: {kayit[1]}")
                print(f"\tSoyadı: {kayit[2]}")
                print(f"\tÜrün Adı: {kayit[3]}")
                print(f"\tTarih: {kayit[4]}")
            secim = input("Hangi kaydı silmek istiyorsunuz? (Numara): ")
            try:
                secim = int(secim)
                if secim < 1 or secim > len(kayitlar):
                    raise ValueError()
            except ValueError:
                print("Geçersiz seçim. Lütfen bir sayı girin.")
                continue
            kayit = kayitlar[secim - 1]
        else:
            kayit = kayitlar[0]

        ad = kayit[1]
        soyad = kayit[2]
        urun_ad = kayit[3]

        # Kaydı veri çerçevesinden sil
        for i, row in enumerate(ws.iter_rows(min_row=2, min_col=1, values_only=True)):
            if row[0] == tc_no and row[1] == ad and row[2] == soyad and row[3] == urun_ad:
                ws.delete_rows(i + 2, 1)
                break

        # Dosyaya kaydet
        wb.save(dosya_yolu)
        print("Kayıt başarıyla silindi.")
        print(f"{ad} {soyad} isimli kişinin {urun_ad} ürünü için kaydı silindi.")
        os.system('cls' if os.name == 'nt' else 'clear')



def kayitlar():
    wb = load_workbook(dosya_yolu)
    ws = wb.active

    header = ["TC Kimlik No", "Ad", "Soyad", "Ürün Adı", "Tarih"]
    kayitlar = [header]

    for row in ws.iter_rows(min_row=2, values_only=True):
        kayitlar.append(row)

    os.system("cls")
    print("{:<15} {:<15} {:<15} {:<15} {:<25}".format(*header))
    print("-" * 80)
    for row in kayitlar[1:]:
        print("{:<15} {:<15} {:<15} {:<15} {:<25}".format(*row))




# Verilerin kaydedileceği dosya yolu
dosya_yolu = "depo_programı.xlsx"

while True:
    print("""
╔══════════════════════════════════════════════╗
║           DEPO OTOMASYONU PROGRAMI            ║
║                 Ana Menü                     ║
╠══════════════════════════════════════════════╣
║ Lütfen yapmak istediğiniz işlemi seçin:      ║
║                                              ║
║ 1 - Kayıt Ekle                               ║
║ 2 - Kayıtları Listele                        ║
║ 3 - Kayıt Sil                                ║
║ Q - Çıkış                                    ║
╚══════════════════════════════════════════════╝
    """)

    secim = input("Seçiminiz: ")

    if secim == "1":
        os.system("cls")
        kayit_ekle()
    elif secim == "2":
        os.system("cls")
        kayitlar()
    elif secim == "3":
        os.system("cls")
        kayit_sil()
    elif secim.lower() == "q":
        break
    else:
        os.system("cls")
        print("Geçersiz seçim. Lütfen tekrar deneyin.")

